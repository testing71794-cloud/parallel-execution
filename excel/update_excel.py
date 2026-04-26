"""
Thread-safe updates to final_execution_report.xlsx.
Fresh run: prime_workbook() creates a new file with headers only; append_result_row() adds rows.
"""
from __future__ import annotations

import logging
import threading
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

logger = logging.getLogger("orch.excel")

# Matches stakeholder report (column "Status", not "Test Status")
HEADERS = [
    "Timestamp",
    "Device Name",
    "Flow Name",
    "Status",
    "Failure Message",
    "AI Analysis",
    "Duration",
]

_file_locks: dict[str, threading.Lock] = {}
_meta_lock = threading.Lock()


def _lock_for(path: Path) -> threading.Lock:
    key = str(path.resolve())
    with _meta_lock:
        if key not in _file_locks:
            _file_locks[key] = threading.Lock()
        return _file_locks[key]


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    """Accept legacy key "Test Status" as "Status"."""
    out = dict(row)
    if "Status" not in out and "Test Status" in out:
        out["Status"] = out.get("Test Status", "")
    return out


def prime_workbook(excel_path: Path, *, file_lock: threading.Lock | None = None) -> None:
    """
    Create a brand-new workbook with header row only (call once per run, after cleanup).
    """
    excel_path = excel_path.resolve()
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    lock = file_lock or _lock_for(excel_path)
    with lock:
        if excel_path.is_file():
            excel_path.unlink()
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(HEADERS)
        for c in ws[1]:
            c.font = Font(bold=True)
        wb.save(excel_path)
        logger.info("Primed new Excel workbook: %s", excel_path)


def append_result_row(
    excel_path: Path,
    row: dict[str, Any],
    *,
    file_lock: threading.Lock | None = None,
) -> None:
    """
    Append one data row. Workbook must already exist (use prime_workbook first).
    """
    excel_path = excel_path.resolve()
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    row = _normalize_row(row)

    lock = file_lock or _lock_for(excel_path)
    values = [row.get(h, "") for h in HEADERS]

    with lock:
        if not excel_path.is_file():
            logger.warning("Excel missing; creating with headers (unexpected — run prime_workbook)")
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            ws.append(HEADERS)
            for c in ws[1]:
                c.font = Font(bold=True)
        else:
            wb = load_workbook(excel_path)
            ws = wb.active
            if ws.max_row == 0:
                ws.append(HEADERS)
                for c in ws[1]:
                    c.font = Font(bold=True)

        ws.append(values)
        wb.save(excel_path)
        logger.info(
            "Excel updated | flow=%s | device=%s | status=%s",
            row.get("Flow Name"),
            row.get("Device Name"),
            row.get("Status"),
        )


def finalize_workbook(excel_path: Path, *, file_lock: threading.Lock | None = None) -> None:
    excel_path = excel_path.resolve()
    if not excel_path.is_file():
        logger.warning("finalize: file missing %s", excel_path)
        return
    lock = file_lock or _lock_for(excel_path)
    with lock:
        wb = load_workbook(excel_path)
        wb.save(excel_path)
    logger.info("Excel finalized: %s", excel_path)
