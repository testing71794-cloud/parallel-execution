#!/usr/bin/env python3
"""Merge per-device ATP JSON into reports/final_report.xlsx (uses scripts.failure_row_analysis)."""
from __future__ import annotations

import json
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))
if str(REPO / "scripts") not in sys.path:
    sys.path.insert(0, str(REPO / "scripts"))
from failure_row_analysis import analyze_failure_for_row
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from utils.device_utils import get_device_name

A = REPO / "automation_with_testcases"
OUT = A / "reports" / "final_report.xlsx"
RBASE = A / "results"
COLS = [
    "Suite",
    "Flow Name",
    "Device Name",
    "Device ID",
    "Status",
    "Exit Code",
    "Retry Count",
    "Failure Step",
    "Error Message",
    "AI Failure Summary",
    "Root Cause Category",
    "Suggested Fix",
    "AI Confidence",
    "Analysis Source",
    "Log Path",
    "Screenshot Path",
    "Timestamp",
]
PASS = PatternFill(fill_type="solid", fgColor="C6EFCE")
FAIL = PatternFill(fill_type="solid", fgColor="FFC7CE")
GRAY = PatternFill(fill_type="solid", fgColor="D9D9D9")
HDR = PatternFill(fill_type="solid", fgColor="B4C7E7")
use_or = os.environ.get("EXCEL_AI_OPENROUTER", "0").lower() in (
    "1",
    "true",
    "yes",
) and (
    (REPO / "build-summary" / "ai_status.txt").is_file()
    and "AI_STATUS=AVAILABLE" in (REPO / "build-summary" / "ai_status.txt").read_text(
        encoding="utf-8", errors="ignore"
    )
)


def load_rows() -> list[dict]:
    out: list[dict] = []
    if not RBASE.is_dir():
        RBASE.mkdir(parents=True, exist_ok=True)
    for sub in sorted(p for p in RBASE.iterdir() if p.is_dir()):
        js = sub / "atp_device_status.json"
        if not js.is_file():
            continue
        try:
            d = json.loads(js.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        did = str(d.get("deviceId", ""))
        seg = str(d.get("deviceSeg", sub.name))
        dname = str(d.get("deviceName", "")) or get_device_name(did)
        logp = str(
            d.get("logPath", (A / "logs" / seg / "last_maestro.log").as_posix())
        ).replace("/", "\\")
        mlog = Path(logp)
        if not mlog.is_file():
            mlog = A / "logs" / seg / "last_maestro.log"
        logp = str(mlog.resolve()) if mlog.is_file() else str(logp)
        try:
            ecn = int(d.get("maestroExit", -1))
        except (TypeError, ValueError):
            ecn = -1
        st = "PASS" if ecn == 0 else "FAIL"
        an = analyze_failure_for_row(
            logp if st == "FAIL" else None,
            status="FAIL" if st == "FAIL" else "PASS",
            use_openrouter=use_or and st == "FAIL",
        )
        if st == "PASS":
            an = {
                "failure_step": "",
                "error_message": "",
                "ai_failure_summary": "—",
                "root_cause_category": "—",
                "suggested_fix": "—",
                "ai_confidence": 1.0,
                "analysis_source": "N/A",
            }
        out.append(
            {
                "Suite": "atp",
                "Flow Name": str(d.get("flow", "signup_atp_smoke")),
                "Device Name": dname,
                "Device ID": did,
                "Status": st,
                "Exit Code": str(ecn),
                "Retry Count": "0",
                "Failure Step": (an.get("failure_step", "") or "")[:2000],
                "Error Message": (an.get("error_message", "") or "")[:2000],
                "AI Failure Summary": (an.get("ai_failure_summary", "") or "See log")[:2000],
                "Root Cause Category": (an.get("root_cause_category", "Unknown") or "Unknown")[
                    :120
                ],
                "Suggested Fix": (an.get("suggested_fix", "") or "Review log.")[:2000],
                "AI Confidence": float(an.get("ai_confidence", 0.65) or 0.65),
                "Analysis Source": str(
                    an.get("analysis_source", "Rule-based fallback")
                )[:60],
                "Log Path": logp,
                "Screenshot Path": str((A / "screenshots" / seg).resolve()),
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )
    if not out:
        an = analyze_failure_for_row(
            None, status="FAIL", use_openrouter=False
        )
        out.append(
            {
                "Suite": "atp",
                "Flow Name": "—",
                "Device Name": "",
                "Device ID": "",
                "Status": "FAIL",
                "Exit Code": "1",
                "Retry Count": "0",
                "Failure Step": an.get("failure_step", "No ATP results"),
                "Error Message": "No ATP execution result found",
                "AI Failure Summary": an.get("ai_failure_summary", "No per-device status JSON")[
                    :2000
                ],
                "Root Cause Category": "Config/Setup Issue",
                "Suggested Fix": "Check ADB, Maestro, and atp_orchestrator.log",
                "AI Confidence": 0.5,
                "Analysis Source": "Rule-based fallback",
                "Log Path": str((A / "logs" / "atp_orchestrator.log").resolve()),
                "Screenshot Path": str((A / "screenshots").resolve()),
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )
    return out


def _autosize(ws, mx=50):
    for c in range(1, (ws.max_column or 1) + 1):
        w = 10
        for r in range(1, min(ws.max_row or 1, 150) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                w = min(max(w, len(str(v)) + 1), mx)
        ws.column_dimensions[get_column_letter(c)].width = w


def main() -> int:
    rows = load_rows()
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Summary"
    t, p, f = len(rows), sum(1 for r in rows if r["Status"] == "PASS"), sum(
        1 for r in rows if r["Status"] == "FAIL"
    )
    ws0["A1"] = "ATP merged report (all devices)"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A2"], ws0["B2"] = "Total", str(t)
    ws0["A3"], ws0["B3"] = "Passed", str(p)
    ws0["A4"], ws0["B4"] = "Failed", str(f)
    wr = wb.create_sheet("Raw Results")
    wr.append(COLS)
    for c in wr[1]:
        c.fill = HDR
    for r in rows:
        wr.append([r.get(c, "") for c in COLS])
    for i in range(2, wr.max_row + 1):
        st = str(wr.cell(i, 5).value or "").upper()
        wr.cell(i, 5).fill = (
            {"PASS": PASS, "FAIL": FAIL}.get(st, GRAY)
        )
    _autosize(wr, 50)
    wdev = wb.create_sheet("Device Summary")
    wdev.append(["Device Name", "Device ID", "Total", "Pass", "Fail"])
    for c in wdev[1]:
        c.fill = HDR
    by = defaultdict(list)
    for r in rows:
        by[str(r.get("Device ID", ""))].append(r)
    for did, arr in sorted(by.items()):
        wdev.append(
            [
                arr[0].get("Device Name", ""),
                did,
                len(arr),
                sum(1 for x in arr if x.get("Status") == "PASS"),
                sum(1 for x in arr if x.get("Status") != "PASS"),
            ]
        )
    wdf = wb.create_sheet("Failure Details")
    wdf.append(COLS)
    for c in wdf[1]:
        c.fill = HDR
    for r in rows:
        if (r.get("Status") or "").upper() == "FAIL":
            wdf.append([r.get(c, "") for c in COLS])
    _autosize(wdf, 50)
    _autosize(wdev, 40)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT} rows={len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
