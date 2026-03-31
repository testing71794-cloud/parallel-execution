from __future__ import annotations

import csv
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WORKSPACE = Path(r"C:\JenkinsAgent\workspace\Kodak-smile-automation")
REPORTS_DIR = WORKSPACE / "reports"
BUILD_SUMMARY_DIR = WORKSPACE / "build-summary"

THIN = Side(style="thin", color="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
PASS_FILL = PatternFill("solid", fgColor="E2F0D9")
FAIL_FILL = PatternFill("solid", fgColor="FCE4D6")
TITLE_FONT = Font(size=15, bold=True)
HEADER_FONT = Font(bold=True)
LINK_FONT = Font(color="0563C1", underline="single")


def safe_int(value: str) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return 0


def load_rows(csv_path: Path) -> list[dict[str, str]]:
    if not csv_path.exists():
        return []
    with csv_path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            normalized = {str(k).strip().lower(): (v or "").strip() for k, v in row.items() if k is not None}
            rows.append(
                {
                    "suite": normalized.get("suite", ""),
                    "flow": normalized.get("flow", ""),
                    "device": normalized.get("device", ""),
                    "status": normalized.get("status", ""),
                    "exit_code": normalized.get("exit_code", normalized.get("exit", "0")),
                    "reason": normalized.get("reason", ""),
                    "log_file": normalized.get("log_file", normalized.get("log", "")),
                }
            )
        return rows


def apply_table_style(ws, start_row: int, end_row: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(vertical="top")
    for cell in ws[start_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, widths: dict[int, int] | None = None) -> None:
    widths = widths or {}
    for col_idx, forced_width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = forced_width
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        if col_idx in widths:
            continue
        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)


def build_suite_sheet(wb: Workbook, suite_name: str, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet(title=(suite_name[:28] + " Summary")[:31])
    total = len(rows)
    passed = sum(1 for r in rows if r["status"].upper() == "PASS")
    failed = total - passed

    ws["A1"] = f"{suite_name.title()} Execution Summary"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Total Results"
    ws["B3"] = total
    ws["A4"] = "Passed"
    ws["B4"] = passed
    ws["A5"] = "Failed"
    ws["B5"] = failed

    headers = ["Suite", "Flow", "Device", "Status", "Exit Code", "Reason", "Log"]
    header_row = 8
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=header)

    data_row = header_row + 1
    for row in rows:
        ws.cell(row=data_row, column=1, value=row["suite"])
        ws.cell(row=data_row, column=2, value=row["flow"])
        ws.cell(row=data_row, column=3, value=row["device"])
        status_cell = ws.cell(row=data_row, column=4, value=row["status"])
        ws.cell(row=data_row, column=5, value=safe_int(row["exit_code"]))
        ws.cell(row=data_row, column=6, value=row["reason"])

        log_path = row["log_file"]
        log_cell = ws.cell(row=data_row, column=7)
        if log_path:
            log_cell.value = "View Log"
            log_cell.hyperlink = log_path
            log_cell.font = LINK_FONT
        else:
            log_cell.value = ""

        if row["status"].upper() == "PASS":
            status_cell.fill = PASS_FILL
        else:
            status_cell.fill = FAIL_FILL

        data_row += 1

    apply_table_style(ws, header_row, max(header_row, data_row - 1), len(headers))
    ws.freeze_panes = "A9"
    autosize(ws, widths={7: 45})


def build_summary_sheet(wb: Workbook, all_rows: list[dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "Summary"

    total = len(all_rows)
    passed = sum(1 for r in all_rows if r["status"].upper() == "PASS")
    failed = total - passed

    ws["A1"] = "Kodak Smile Execution Summary"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Total Results"
    ws["B3"] = total
    ws["A4"] = "Passed"
    ws["B4"] = passed
    ws["A5"] = "Failed"
    ws["B5"] = failed

    headers = ["Suite", "Flow", "Device", "Status", "Exit Code", "Reason", "Log"]
    header_row = 8
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx, value=header)

    row_no = header_row + 1
    for row in all_rows:
        ws.cell(row=row_no, column=1, value=row["suite"])
        ws.cell(row=row_no, column=2, value=row["flow"])
        ws.cell(row=row_no, column=3, value=row["device"])
        status_cell = ws.cell(row=row_no, column=4, value=row["status"])
        ws.cell(row=row_no, column=5, value=safe_int(row["exit_code"]))
        ws.cell(row=row_no, column=6, value=row["reason"])

        log_path = row["log_file"]
        log_cell = ws.cell(row=row_no, column=7)
        if log_path:
            log_cell.value = "View Log"
            log_cell.hyperlink = log_path
            log_cell.font = LINK_FONT

        if row["status"].upper() == "PASS":
            status_cell.fill = PASS_FILL
        else:
            status_cell.fill = FAIL_FILL
        row_no += 1

    apply_table_style(ws, header_row, max(header_row, row_no - 1), len(headers))
    ws.freeze_panes = "A9"
    autosize(ws, widths={7: 45})

    suite_sheet = wb.create_sheet(title="Suite Summary")
    suite_headers = ["Suite", "Total", "Passed", "Failed"]
    for idx, header in enumerate(suite_headers, start=1):
        suite_sheet.cell(row=1, column=idx, value=header)

    by_suite: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in all_rows:
        by_suite[row["suite"]].append(row)

    current_row = 2
    for suite_name, rows in sorted(by_suite.items()):
        total_suite = len(rows)
        pass_suite = sum(1 for r in rows if r["status"].upper() == "PASS")
        fail_suite = total_suite - pass_suite
        suite_sheet.cell(row=current_row, column=1, value=suite_name)
        suite_sheet.cell(row=current_row, column=2, value=total_suite)
        suite_sheet.cell(row=current_row, column=3, value=pass_suite)
        suite_sheet.cell(row=current_row, column=4, value=fail_suite)
        current_row += 1

    apply_table_style(suite_sheet, 1, max(1, current_row - 1), len(suite_headers))
    autosize(suite_sheet)


def generate_failed_summary(all_rows: list[dict[str, str]]) -> str:
    failed_rows = [r for r in all_rows if r["status"].upper() != "PASS"]
    if not failed_rows:
        return "All flows passed."
    lines = []
    for row in failed_rows:
        lines.append(
            f"- Suite={row['suite']} | Flow={row['flow']} | Device={row['device']} | Exit={row['exit_code']} | Reason={row['reason']}"
        )
    return "\n".join(lines)


def main() -> int:
    BUILD_SUMMARY_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    suite_csvs = list(REPORTS_DIR.glob("*/all_results.csv"))
    all_rows: list[dict[str, str]] = []
    rows_by_suite: dict[str, list[dict[str, str]]] = {}

    for csv_path in suite_csvs:
        suite_name = csv_path.parent.name
        rows = load_rows(csv_path)
        if rows:
            rows_by_suite[suite_name] = rows
            all_rows.extend(rows)

    if not all_rows:
        print("No suite result CSV files found.")
        return 1

    wb = Workbook()
    build_summary_sheet(wb, all_rows)
    for suite_name, rows in sorted(rows_by_suite.items()):
        build_suite_sheet(wb, suite_name, rows)

    output_path = BUILD_SUMMARY_DIR / "final_execution_report.xlsx"
    wb.save(output_path)

    failed_text = generate_failed_summary(all_rows)
    (BUILD_SUMMARY_DIR / "failed_summary.txt").write_text(failed_text, encoding="utf-8")

    html_lines = [
        "<html><body>",
        "<h2>Kodak Smile Execution Summary</h2>",
        f"<p><b>Total:</b> {len(all_rows)} &nbsp; <b>Passed:</b> {sum(1 for r in all_rows if r['status'].upper() == 'PASS')} &nbsp; <b>Failed:</b> {sum(1 for r in all_rows if r['status'].upper() != 'PASS')}</p>",
        "<h3>Failed Flows</h3>",
        "<pre>",
        failed_text,
        "</pre>",
        "</body></html>",
    ]
    (BUILD_SUMMARY_DIR / "summary.html").write_text("\n".join(html_lines), encoding="utf-8")

    print(f"Generated report: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
