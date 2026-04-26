#!/usr/bin/env python3
"""
Read final_execution_report.xlsx (orchestrator format) and emit pass/fail/total counts.
Writes a small text file for Jenkins / operators.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python scripts/summarize_final_excel.py <final_execution_report.xlsx> [counts_out.txt]")
        return 1

    xlsx = Path(sys.argv[1]).resolve()
    out_txt = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else xlsx.parent / "build-summary" / "execution_counts.txt"

    if not xlsx.is_file():
        print(f"[WARN] Excel not found: {xlsx}")
        out_txt.parent.mkdir(parents=True, exist_ok=True)
        out_txt.write_text("Total=0\nPassed=0\nFailed=0\n", encoding="utf-8")
        return 0

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("Empty workbook")
        return 1

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    try:
        status_idx = headers.index("Status")
    except ValueError:
        print("No 'Status' column in first row", headers)
        return 1

    total = 0
    passed = 0
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        total += 1
        st = ""
        if status_idx < len(row) and row[status_idx] is not None:
            st = str(row[status_idx]).strip().upper()
        if st == "PASS":
            passed += 1

    failed = total - passed
    lines = [
        f"Total={total}",
        f"Passed={passed}",
        f"Failed={failed}",
        "",
        f"Source={xlsx}",
    ]
    out_txt.parent.mkdir(parents=True, exist_ok=True)
    out_txt.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"[INFO] Summary: Total={total} Passed={passed} Failed={failed} -> {out_txt}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
