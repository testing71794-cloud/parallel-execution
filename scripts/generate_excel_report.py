#!/usr/bin/env python3
"""
Per-suite Excel + merged build-summary/final_execution_report.xlsx.
Always writes output files; uses AI / rule-based columns for failed & flaky rows.
"""
from __future__ import annotations

import csv
import os
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))
if str(REPO / "scripts") not in sys.path:
    sys.path.insert(0, str(REPO / "scripts"))

from failure_row_analysis import analyze_failure_for_row
from utils.device_utils import get_device_name

PASS_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
FLAKY_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="B4C7E7")
GRAY_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")

COLS = [
    "Suite",
    "Flow Name",
    "Device Name",
    "Device ID",
    "Status",
    "AI Status",
    "Model Used",
    "Exit Code",
    "Retry Count",
    "Failure Step",
    "Error Message",
    "AI Failure Summary",
    "AI Analyses",
    "Root Cause Category",
    "Suggested Fix",
    "AI Confidence",
    "Analysis Source",
    "Log Path",
    "Screenshot Path",
    "Timestamp",
    "AI Analysis",
]

_SCREEN_DEFAULT = str((REPO / ".maestro" / "screenshots").resolve())

# Flow-only view for final report / email (no suite-level rows without a flow)
FLOW_REPORT_HEADERS: tuple[str, ...] = (
    "Suite",
    "Flow",
    "Device",
    "Device ID",
    "Status",
    "Exit Code",
    "AI Analysis",
)


def _augment_merged_row(rowd: dict) -> None:
    """Refresh Device Name from ADB and AI Analysis for rows loaded from an older xlsx merge."""
    did = str(rowd.get("Device ID", "") or "").strip()
    if did:
        rowd["Device Name"] = get_device_name(did)
    ai = (
        str(rowd.get("AI Analysis", "") or "").strip()
        or str(rowd.get("AI Analyses", "") or "").strip()
        or str(rowd.get("AI Failure Summary", "") or "").strip()
        or "—"
    )
    rowd["AI Analysis"] = ai


def _write_flow_report_sheet(wb: Workbook, all_rows: list[dict]) -> None:
    w = wb.create_sheet("Flow Report", 1)
    w.append(list(FLOW_REPORT_HEADERS))
    for c in w[1]:
        c.fill = HEADER_FILL
        c.font = Font(bold=True)
    for r in all_rows:
        flow = str(r.get("Flow Name", "") or "").strip()
        if not flow:
            continue
        suite_s = str(r.get("Suite", "") or "").strip()
        if not suite_s:
            continue
        did = str(r.get("Device ID", "") or "").strip()
        dev = str(r.get("Device Name", "") or "").strip() or (get_device_name(did) if did else "")
        st = str(r.get("Status", "") or "").strip()
        ex = str(r.get("Exit Code", "") or "").strip() or "0"
        ai = (
            str(r.get("AI Analysis", "") or "").strip()
            or str(r.get("AI Analyses", "") or "").strip()
            or str(r.get("AI Failure Summary", "") or "").strip()
            or "—"
        )
        w.append([suite_s, flow, dev, did, st, ex, ai])
    _autosize(w, 60)


def parse_status_file(file_path: Path) -> dict:
    data: dict = {
        "suite": "",
        "flow": "",
        "device": "",
        "device_id": "",
        "device_name": "",
        "status": "",
        "log": "",
        "exit_code": "",
        "reason": "",
        "log_file": "",
        "log_path": "",
        "first_log_path": "",
        "retry_count": "0",
        "timestamp": "",
        "file_name": file_path.name,
    }
    try:
        for line in file_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            data[key.strip().lower()] = value.strip()
    except Exception as exc:
        data["status"] = "PARSE_ERROR"
        data["log"] = f"Could not parse file: {exc}"

    if not data.get("flow"):
        stem = file_path.stem
        parts = stem.split("__")
        if len(parts) >= 3:
            data["suite"] = data.get("suite") or parts[0]
            data["flow"] = data.get("flow") or parts[1]
            data["device"] = data.get("device") or parts[2]
    data["status"] = (data.get("status") or "UNKNOWN").upper()
    return data


def _device_id(row: dict) -> str:
    return (row.get("device_id") or row.get("device") or "").strip()


def _display_device_name_for_report(row: dict) -> str:
    """
    Device column: always prefer ADB "Brand Model" for a serial, so Excel shows Samsung Galaxy
    S21, not RZCT... . Falls back to status device_name if no id; then get_device_name on id.
    """
    did = _device_id(row)
    if did:
        return get_device_name(did)
    return (row.get("device_name") or "").strip()


def _log_path(row: dict) -> str:
    for k in ("log_path", "first_log_path", "log_file", "log"):
        v = (row.get(k) or "").strip()
        if v and v != "PARSE_ERROR":
            return str(Path(v).resolve()) if ("/" in v or "\\" in v or ":" in v) else v
    return ""


def load_results(status_dir: Path, suite_name: str) -> list[dict]:
    if not status_dir.exists():
        return []
    results = []
    for file_path in sorted(status_dir.glob("*.txt")):
        row = parse_status_file(file_path)
        row_suite = (row.get("suite") or "").strip().lower()
        if suite_name and row_suite and row_suite != suite_name.lower():
            continue
        if row.get("status") == "RUNNING":
            continue
        results.append(row)
    return results


def _rows_to_raw_dicts(
    results: list[dict], suite_label: str, use_or: bool
) -> list[dict]:
    out: list[dict] = []
    for row in results:
        did = _device_id(row)
        dname = _display_device_name_for_report(row)
        st = (row.get("status") or "UNKNOWN").upper()
        logp = _log_path(row)
        ec = (row.get("exit_code") or "").strip() or "0"
        flow = (row.get("flow") or "").strip()
        if st in ("FAIL", "FLAKY") or (st != "PASS" and st not in ("UNKNOWN", "RUNNING")):
            st_for_ai = st if st in ("FAIL", "FLAKY") else "FAIL"
            an = analyze_failure_for_row(
                logp if logp else None, status=st_for_ai, use_openrouter=use_or
            )
        else:
            an = analyze_failure_for_row(
                None, status="PASS", use_openrouter=False
            )
        ai_raw = (an.get("ai_failure_summary") or "").strip()
        ai_one = (ai_raw[:2000] if ai_raw else "—")
        out.append(
            {
                "Suite": suite_label,
                "Flow Name": flow,
                "Device Name": dname,
                "Device ID": did,
                "Status": st,
                "AI Status": str(
                    an.get("ai_status") or "NOT_CHECKED"
                )[:40],
                "Model Used": str(
                    an.get("model_used") or "—"
                )[:120],
                "Exit Code": str(ec),
                "Retry Count": (row.get("retry_count") or "0")[:8],
                "Failure Step": (an.get("failure_step") or "")[:2000],
                "Error Message": (an.get("error_message") or "")[:2000],
                "AI Failure Summary": (an.get("ai_failure_summary") or "—")[:2000],
                "AI Analyses": (an.get("ai_failure_summary") or "—")[:2000],
                "Root Cause Category": (an.get("root_cause_category") or "—")[:120],
                "Suggested Fix": (an.get("suggested_fix") or "—")[:2000],
                "AI Confidence": float(an.get("ai_confidence", 0.65) or 0.65),
                "Analysis Source": (
                    an.get("analysis_source")
                    or "Rule-based fallback"
                )[:60],
                "Log Path": logp,
                "Screenshot Path": _SCREEN_DEFAULT,
                "Timestamp": (row.get("timestamp") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"))[
                    :32
                ],
                "AI Analysis": ai_one,
            }
        )
    return out


def _fill_raw(ws, rows: list[dict]) -> None:
    ws.append(COLS)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in COLS])
    for i in range(2, ws.max_row + 1):
        st = str(ws.cell(i, 5).value or "").upper()
        cell = ws.cell(i, 5)
        cell.fill = {
            "PASS": PASS_FILL,
            "FAIL": FAIL_FILL,
            "FLAKY": FLAKY_FILL,
        }.get(st, GRAY_FILL)


def _autosize(ws, mx: int = 55) -> None:
    for c in range(1, (ws.max_column or 1) + 1):
        w = 10
        for r in range(1, min(ws.max_row or 1, 200) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                w = min(max(w, len(str(v)) + 1), mx)
        ws.column_dimensions[get_column_letter(c)].width = w


def _merge_build_summary(
    suite_key: str, new_rows: list[dict], build_summary: Path
) -> None:
    build_summary.mkdir(parents=True, exist_ok=True)
    final = build_summary / "final_execution_report.xlsx"
    by_suites: dict[str, list[dict]] = {suite_key.lower(): new_rows}
    if final.is_file():
        try:
            wb_o = load_workbook(final)
            if "Raw Results" in wb_o.sheetnames:
                ws = wb_o["Raw Results"]
                h = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
                for r in range(2, (ws.max_row or 1) + 1):
                    d: dict = {}
                    for ci, name in enumerate(h, start=1):
                        d[name] = ws.cell(r, ci).value
                    su = str(d.get("Suite") or "").strip().lower()
                    if not su or su == suite_key.lower():
                        continue
                    if su not in by_suites:
                        by_suites[su] = []
                    rowd: dict = {c: d.get(c, "") for c in COLS}
                    if (not str(rowd.get("AI Analyses", "")).strip()) and str(
                        rowd.get("AI Failure Summary", "")
                    ).strip():
                        rowd["AI Analyses"] = rowd.get("AI Failure Summary", "")
                    _augment_merged_row(rowd)
                    by_suites[su].append(rowd)
        except Exception as exc:
            print(f"Note: could not merge prior final_execution_report.xlsx: {exc}")
    all_rows: list[dict] = []
    for _k in sorted(by_suites.keys()):
        for rowd in by_suites[_k]:
            all_rows.append({c: rowd.get(c, "") for c in COLS})
    wb = Workbook()
    # drop default sheet
    wb.remove(wb.active)
    ws0 = wb.create_sheet("Summary", 0)
    t, p, nf, fl = len(all_rows), 0, 0, 0
    for r in all_rows:
        s = (str(r.get("Status") or "")).upper()
        if s == "PASS":
            p += 1
        elif s == "FLAKY":
            fl += 1
        else:
            nf += 1
    ws0["A1"] = "Kodak Smile — merged execution report"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A2"], ws0["B2"] = "Total rows", str(t)
    ws0["A3"], ws0["B3"] = "Passed", str(p)
    ws0["A4"], ws0["B4"] = "Failed (non-PASS, excl. flaky count below)", str(nf)
    ws0["A5"], ws0["B5"] = "Flaky", str(fl)
    ws0["A6"], ws0["B6"] = "Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _write_flow_report_sheet(wb, all_rows)
    wdev = wb.create_sheet("Device Summary")
    wdev.append(["Device Name", "Device ID", "Total", "Passed", "Failed", "Flaky"])
    by_d: dict[str, list[dict]] = defaultdict(list)
    for r in all_rows:
        by_d[str(r.get("Device ID", ""))].append(r)
    for did, arr in sorted(by_d.items(), key=lambda x: x[0]):
        wdev.append(
            [
                arr[0].get("Device Name", ""),
                did,
                len(arr),
                sum(1 for x in arr if (str(x.get("Status") or "")).upper() == "PASS"),
                sum(
                    1
                    for x in arr
                    if (str(x.get("Status") or "")).upper() not in ("PASS", "FLAKY")
                ),
                sum(1 for x in arr if (str(x.get("Status") or "")).upper() == "FLAKY"),
            ]
        )
    wdf = wb.create_sheet("Failure Details")
    wdf.append(COLS)
    for c in wdf[1]:
        c.fill = HEADER_FILL
    for r in all_rows:
        st = (str(r.get("Status") or "")).upper()
        if st in ("FAIL", "FLAKY", "PARSE_ERROR", "UNKNOWN"):
            wdf.append([r.get(c, "") for c in COLS])
    wr = wb.create_sheet("Raw Results")
    _fill_raw(wr, all_rows)
    _autosize(wr, 50)
    _autosize(wdf, 50)
    _autosize(wdev, 40)
    _autosize(ws0, 40)
    wb.save(final)


def build_workbook(
    results: list[dict],
    output_file: Path,
    suite_name: str,
    suite_label: str,
    use_or: bool,
) -> list[dict]:
    raw = _rows_to_raw_dicts(results, suite_label, use_or)
    wb = Workbook()
    t, p, f, fl = len(raw), 0, 0, 0
    for r in raw:
        s = (r.get("Status") or "").upper()
        if s == "PASS":
            p += 1
        elif s == "FLAKY":
            fl += 1
        else:
            f += 1
    ws0 = wb.active
    ws0.title = "Summary"
    ws0["A1"] = f"Suite: {suite_label}"
    ws0["A1"].font = Font(bold=True, size=14)
    ws0["A2"], ws0["B2"] = "Total", str(t)
    ws0["A3"], ws0["B3"] = "Passed", str(p)
    ws0["A4"], ws0["B4"] = "Failed", str(f)
    ws0["A5"], ws0["B5"] = "Flaky", str(fl)
    ws0["A6"] = "Generated"
    ws0["B6"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    _write_flow_report_sheet(wb, raw)
    wdev = wb.create_sheet("Device Summary")
    wdev.append(["Device Name", "Device ID", "Total", "Passed", "Failed", "Flaky"])
    by_d: dict[str, list[dict]] = defaultdict(list)
    for r in raw:
        by_d[str(r.get("Device ID", ""))].append(r)
    for did, arr in sorted(by_d.items(), key=lambda x: x[0]):
        wdev.append(
            [
                arr[0].get("Device Name", ""),
                did,
                len(arr),
                sum(1 for x in arr if (x.get("Status") or "").upper() == "PASS"),
                sum(
                    1
                    for x in arr
                    if (x.get("Status") or "").upper() not in ("PASS", "FLAKY")
                ),
                sum(1 for x in arr if (x.get("Status") or "").upper() == "FLAKY"),
            ]
        )
    wdf = wb.create_sheet("Failure Details")
    wdf.append(COLS)
    for c in wdf[1]:
        c.fill = HEADER_FILL
    for r in raw:
        if (r.get("Status") or "").upper() != "PASS":
            wdf.append([r.get(c, "") for c in COLS])
    wr = wb.create_sheet("Raw Results")
    _fill_raw(wr, raw)
    for sheet in (ws0, wdev, wdf, wr):
        _autosize(sheet, 50)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    return raw


def _ai_use_openrouter() -> bool:
    p = REPO / "build-summary" / "ai_status.txt"
    if p.is_file():
        text = p.read_text(encoding="utf-8", errors="ignore")
        if "AI_STATUS=UNAVAILABLE" in text:
            return False
        if "AI_STATUS=AVAILABLE" in text:
            return True
    return os.environ.get("EXCEL_AI_OPENROUTER", "0").lower() in ("1", "true", "yes")


def write_csv(path: Path, rows: list[dict], only_status: str | None = None):
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = ["suite", "flow", "device", "status", "exit_code", "log", "file_name"]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for row in rows:
            if only_status == "PASS" and row.get("status") != "PASS":
                continue
            if only_status == "FAIL" and row.get("status") == "PASS":
                continue
            w.writerow({h: row.get(h, "") for h in headers})


def main() -> int:
    if len(sys.argv) != 4:
        print("Usage: python scripts/generate_excel_report.py <status_dir> <output_dir> <suite_name>")
        return 1

    status_dir = Path(sys.argv[1]).resolve()
    output_dir = Path(sys.argv[2]).resolve()
    suite_name = sys.argv[3].strip().lower()
    suite_label = sys.argv[3].strip()

    output_dir.mkdir(parents=True, exist_ok=True)
    use_or = _ai_use_openrouter()
    results = load_results(status_dir, suite_name)

    if not results:
        print(
            f"Warning: no completed status results for suite '{suite_name}' in {status_dir} — writing placeholder rows."
        )
        results = [
            {
                "suite": suite_name,
                "flow": "—",
                "device": "",
                "status": "UNKNOWN",
                "exit_code": "0",
                "log": "No status files matched this suite.",
            }
        ]

    raw = build_workbook(
        results, output_dir / "summary.xlsx", suite_name, suite_label, use_or
    )
    write_csv(output_dir / "all_results.csv", results)
    write_csv(output_dir / "failed_results.csv", results, only_status="FAIL")
    write_csv(output_dir / "passed_results.csv", results, only_status="PASS")
    (REPO / "build-summary").mkdir(parents=True, exist_ok=True)
    _merge_build_summary(suite_name, raw, REPO / "build-summary")
    print(f"Report: {output_dir / 'summary.xlsx'} | merged: build-summary/final_execution_report.xlsx | rows={len(raw)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())