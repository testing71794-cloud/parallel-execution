import argparse
import os
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADERS = [
    'Build Number',
    'Suite Type',
    'Flow Order',
    'Flow Name',
    'Device ID',
    'Status',
    'Duration (s)',
    'Failure Reason',
    'XML Report',
    'Log File',
    'Debug Folder',
    'Updated At'
]


def ensure_workbook(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = 'Execution Summary'
    ws.append(HEADERS)

    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    for idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    widths = {
        1: 14, 2: 14, 3: 10, 4: 18, 5: 22, 6: 12,
        7: 14, 8: 40, 9: 38, 10: 38, 11: 40, 12: 22
    }
    for col, width in widths.items():
        ws.column_dimensions[chr(64 + col)].width = width
    return wb, ws


def parse_xml(xml_path: str):
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except Exception as exc:
        return [{
            'status': 'FAIL',
            'duration': '',
            'reason': f'Invalid XML: {exc}'
        }]

    testcases = root.findall('.//testcase')
    if not testcases and root.tag == 'testcase':
        testcases = [root]

    rows = []
    for tc in testcases:
        failure = tc.find('failure')
        error = tc.find('error')
        skipped = tc.find('skipped')
        reason = ''
        status = 'PASS'
        if failure is not None:
            status = 'FAIL'
            reason = (failure.attrib.get('message') or failure.text or '').strip()
        elif error is not None:
            status = 'FAIL'
            reason = (error.attrib.get('message') or error.text or '').strip()
        elif skipped is not None:
            status = 'SKIPPED'
            reason = (skipped.attrib.get('message') or skipped.text or '').strip()

        rows.append({
            'status': status,
            'duration': tc.attrib.get('time', ''),
            'reason': reason[:500]
        })

    if not rows:
        rows.append({
            'status': 'PASS',
            'duration': '',
            'reason': ''
        })
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--results-dir', required=True)
    parser.add_argument('--workbook', required=True)
    parser.add_argument('--suite-type', required=True)
    parser.add_argument('--flow-name', required=True)
    parser.add_argument('--flow-order', type=int, required=True)
    parser.add_argument('--build-number', default=os.environ.get('BUILD_NUMBER', ''))
    args = parser.parse_args()

    wb, ws = ensure_workbook(args.workbook)
    updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    pattern_prefix = f"{args.flow_name}_"
    xml_files = [
        os.path.join(args.results_dir, name)
        for name in os.listdir(args.results_dir)
        if name.startswith(pattern_prefix) and name.endswith('.xml')
    ]
    xml_files.sort()

    if not xml_files:
        ws.append([
            args.build_number,
            args.suite_type,
            args.flow_order,
            args.flow_name,
            'ALL',
            'FAIL',
            '',
            'No XML results generated for this flow',
            '',
            '',
            '',
            updated_at
        ])
    else:
        for xml_path in xml_files:
            file_name = os.path.basename(xml_path)
            device_id = file_name[len(pattern_prefix):-4]
            log_file = os.path.join(args.results_dir, f'{args.flow_name}_{device_id}.log')
            debug_dir = os.path.join(args.results_dir, f'debug_{args.flow_name}_{device_id}')

            parsed_rows = parse_xml(xml_path)
            for item in parsed_rows:
                ws.append([
                    args.build_number,
                    args.suite_type,
                    args.flow_order,
                    args.flow_name,
                    device_id,
                    item['status'],
                    item['duration'],
                    item['reason'],
                    os.path.relpath(xml_path, os.getcwd()),
                    os.path.relpath(log_file, os.getcwd()) if os.path.exists(log_file) else '',
                    os.path.relpath(debug_dir, os.getcwd()) if os.path.exists(debug_dir) else '',
                    updated_at
                ])

    wb.save(args.workbook)
    print(f'Workbook updated: {args.workbook}')


if __name__ == '__main__':
    main()
