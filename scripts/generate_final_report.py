import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

SEPARATOR = "__"


def parse_status_files(status_dir: Path):
    rows = []
    if not status_dir.exists():
        return rows
    for path in sorted(status_dir.glob("*.*")):
        suffix = path.suffix.lower()
        if suffix not in {".pass", ".fail"}:
            continue
        parts = path.stem.split(SEPARATOR)
        if len(parts) < 3:
            continue
        suite = parts[0]
        flow = parts[1]
        device = SEPARATOR.join(parts[2:])
        status = "PASS" if suffix == ".pass" else "FAIL"
        rows.append({
            "suite": suite,
            "flow": flow,
            "device": device,
            "status": status,
            "status_file": str(path),
            "log_path": str(Path("reports") / suite / flow / device / "logs" / f"{suite}__{flow}__{device}.log"),
            "junit_path": str(Path("reports") / suite / flow / device / "junit" / f"{suite}__{flow}__{device}.xml"),
        })
    return rows


def read_ai_notes(root: Path):
    notes = []
    ai_dir = root / "ai-doctor" / "artifacts"
    for name in ["cursor-report.md", "ai-report.json", "maestro_stdout.log", "maestro_stderr.log"]:
        path = ai_dir / name
        if path.exists():
            notes.append(str(path.relative_to(root)))
    return notes


def autosize(ws):
    for col in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 14), 60)


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    autosize(ws)


def main():
    if len(sys.argv) < 4:
        print("Usage: generate_final_report.py <project_root> <status_dir> <output_xlsx>")
        sys.exit(1)

    root = Path(sys.argv[1])
    status_dir = Path(sys.argv[2])
    output_xlsx = Path(sys.argv[3])
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    rows = parse_status_files(status_dir)
    total = len(rows)
    passed = sum(1 for r in rows if r["status"] == "PASS")
    failed = sum(1 for r in rows if r["status"] == "FAIL")
    by_suite = defaultdict(lambda: {"PASS": 0, "FAIL": 0})
    for r in rows:
        by_suite[r["suite"]][r["status"]] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ["Total results", total],
        ["Passed", passed],
        ["Failed", failed],
        ["AI analysis files", ", ".join(read_ai_notes(root)) or "None"],
    ]
    for suite, counts in sorted(by_suite.items()):
        summary_rows.append([f"{suite} passed", counts["PASS"]])
        summary_rows.append([f"{suite} failed", counts["FAIL"]])
    write_sheet(ws, ["Metric", "Value"], summary_rows)

    all_ws = wb.create_sheet("All Results")
    headers = ["suite", "flow", "device", "status", "status_file", "log_path", "junit_path"]
    write_sheet(all_ws, headers, [[r[h] for h in headers] for r in rows] or [["No results", "", "", "", "", "", ""]])

    fail_ws = wb.create_sheet("Failed Flows")
    fail_rows = [r for r in rows if r["status"] == "FAIL"]
    write_sheet(fail_ws, headers, [[r[h] for h in headers] for r in fail_rows] or [["No failed flows", "", "", "", "", "", ""]])

    pass_ws = wb.create_sheet("Passed Flows")
    pass_rows = [r for r in rows if r["status"] == "PASS"]
    write_sheet(pass_ws, headers, [[r[h] for h in headers] for r in pass_rows] or [["No passed flows", "", "", "", "", "", ""]])

    ai_ws = wb.create_sheet("AI Analysis Files")
    ai_files = read_ai_notes(root)
    write_sheet(ai_ws, ["artifact"], [[a] for a in ai_files] or [["No AI analysis artifacts found"]])

    wb.save(output_xlsx)
    print(output_xlsx)


if __name__ == "__main__":
    main()
