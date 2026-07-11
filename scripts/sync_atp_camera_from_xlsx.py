#!/usr/bin/env python3
"""
Refresh ATP TestCase Flows/camera/atp_camera_mapping.json from Kodak Step Print ATP 2026.xlsx.

Looks for a sheet named Camera (or module column Camera). When the workbook is missing,
prints the path to place the file and exits 0 without changing JSON.

Usage:
  python scripts/sync_atp_camera_from_xlsx.py
  python scripts/sync_atp_camera_from_xlsx.py --xlsx "Kodak Step Print ATP 2026.xlsx"
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
OUT_JSON = REPO / "ATP TestCase Flows" / "camera" / "atp_camera_mapping.json"
OUT_CSV = REPO / "ATP TestCase Flows" / "camera" / "atp_camera_mapping.csv"
DEFAULT_XLSX = REPO / "Kodak Step Print ATP 2026.xlsx"


def _find_xlsx(explicit: Path | None) -> Path | None:
    if explicit and explicit.is_file():
        return explicit
    for candidate in (
        DEFAULT_XLSX,
        REPO / "docs" / "Kodak Step Print ATP 2026.xlsx",
    ):
        if candidate.is_file():
            return candidate
    return None


def _norm_id(value: str) -> str:
    return re.sub(r"\s+", "", (value or "").strip())


def _load_camera_rows(xlsx: Path) -> list[dict[str, str]]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "camera":
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip().lower() for h in rows[0]]

    def col(*names: str) -> int | None:
        for n in names:
            key = n.lower()
            if key in headers:
                return headers.index(key)
        return None

    i_id = col("testcaseid", "atp test case id", "test case id", "id")
    i_title = col("testcasetitle", "test name", "title", "test case title")
    i_steps = col("steps", "test steps", "step description")
    i_expected = col("expected", "expected result", "expected results")
    i_count = col("step count", "steps count", "# steps")
    if i_id is None:
        return []

    out: list[dict[str, str]] = []
    for row in rows[1:]:
        if i_id >= len(row):
            continue
        tc_id = _norm_id(str(row[i_id] or ""))
        if not tc_id.startswith("CA"):
            continue
        title = str(row[i_title] or "").strip() if i_title is not None and i_title < len(row) else ""
        steps = str(row[i_steps] or "").strip() if i_steps is not None and i_steps < len(row) else ""
        expected = (
            str(row[i_expected] or "").strip()
            if i_expected is not None and i_expected < len(row)
            else ""
        )
        step_count = ""
        if i_count is not None and i_count < len(row) and row[i_count] is not None:
            step_count = str(row[i_count]).strip()
        out.append(
            {
                "atpTestCaseId": tc_id,
                "yamlFileName": f"{tc_id}.yaml",
                "testName": f"{tc_id} - {title}" if title else tc_id,
                "flowType": "Camera",
                "excelStepCount": int(float(step_count)) if step_count else 0,
                "atpSteps": steps,
                "atpExpected": expected,
            }
        )
    return out


def _write_csv(rows: list[dict]) -> None:
    import csv

    fields = [
        "ATP Test Case ID",
        "YAML File Name",
        "Test Name",
        "Flow Type",
        "Excel Step Count",
        "ATP Steps",
        "ATP Expected",
    ]
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    "ATP Test Case ID": r["atpTestCaseId"],
                    "YAML File Name": r["yamlFileName"],
                    "Test Name": r["testName"],
                    "Flow Type": r["flowType"],
                    "Excel Step Count": r.get("excelStepCount", ""),
                    "ATP Steps": r.get("atpSteps", ""),
                    "ATP Expected": r.get("atpExpected", ""),
                }
            )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=None)
    args = parser.parse_args()
    xlsx = _find_xlsx(args.xlsx)
    if not xlsx:
        print(
            f"[camera-sync] Workbook not found. Place it at:\n  {DEFAULT_XLSX}\n"
            "Keeping existing atp_camera_mapping.json.",
            flush=True,
        )
        return 0
    try:
        rows = _load_camera_rows(xlsx)
    except ImportError:
        print("[camera-sync] openpyxl required: pip install openpyxl", flush=True)
        return 1
    if not rows:
        print(f"[camera-sync] No CA_* rows parsed from {xlsx}", flush=True)
        return 1
    OUT_JSON.write_text(json.dumps(rows, indent=2) + "\n", encoding="utf-8")
    _write_csv(rows)
    print(f"[camera-sync] Wrote {len(rows)} cases to {OUT_JSON.relative_to(REPO)}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
