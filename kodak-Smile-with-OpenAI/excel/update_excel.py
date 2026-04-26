"""
Thread-safe Excel for parallel orchestration.
Each run: cleanup + prime_workbook() → fresh headers only → append rows for this run only.
"""
from __future__ import annotations

import logging
import threading
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

logger = logging.getLogger("orch.excel")

HEADERS = [
    "Timestamp",
    "Suite",
    "Flow",
    "Device",
    "Status",
    "Exit Code",
    "Log Path",
    "AI Analysis",
]

_file_locks: dict[str, threading.Lock] = {}
_meta_lock = threading.Lock()


def _lock_for(path: Path) -> threading.Lock:
    key = str(path.resolve())
    with _meta_lock:
        if key not in _file_locks:
            _file_locks[key] = threading.Lock()
        return _file_locks[key]


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    """Map legacy column names from older orchestrator versions."""
    out = dict(row)
    if "Flow" not in out and "Flow Name" in out:
        out["Flow"] = out.get("Flow Name", "")
    if "Device" not in out and "Device Name" in out:
        out["Device"] = out.get("Device Name", "")
    if "Status" not in out and "Test Status" in out:
        out["Status"] = out.get("Test Status", "")
    if "Log Path" not in out and "Log" in out:
        out["Log Path"] = out.get("Log", "")
    if "Log Path" not in out and "Failure Message" in out:
        pass
    if "Exit Code" not in out:
        out.setdefault("Exit Code", "")
    if "Suite" not in out:
        out.setdefault("Suite", "")
    if "AI Analysis" not in out:
        out.setdefault("AI Analysis", "")
    out.setdefault("Log Path", "")
    return out


def prime_workbook(excel_path: Path, *, file_lock: threading.Lock | None = None) -> None:
    """New workbook: headers only (after cleanup, no prior rows)."""
    excel_path = excel_path.resolve()
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    lock = file_lock or _lock_for(excel_path)
    with lock:
        if excel_path.is_file():
            try:
                excel_path.unlink()
            except OSError as e:
                logger.warning("Could not delete existing Excel %s: %s", excel_path, e)
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(HEADERS)
        for c in ws[1]:
            c.font = Font(bold=True)
        wb.save(excel_path)
        logger.info("Primed new Excel: %s", excel_path)


def append_result_row(
    excel_path: Path,
    row: dict[str, Any],
    *,
    file_lock: threading.Lock | None = None,
) -> None:
    excel_path = excel_path.resolve()
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    row = _normalize_row(row)

    lock = file_lock or _lock_for(excel_path)
    values = [row.get(h, "") for h in HEADERS]

    with lock:
        if not excel_path.is_file():
            logger.warning("Excel missing; recreating headers")
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            ws.append(HEADERS)
            for c in ws[1]:
                c.font = Font(bold=True)
        else:
            wb = load_workbook(excel_path)
            ws = wb.active
            if ws.max_row == 0:
                ws.append(HEADERS)
                for c in ws[1]:
                    c.font = Font(bold=True)

        ws.append(values)
        wb.save(excel_path)
        logger.info(
            "Excel updated | flow=%s | device=%s | status=%s",
            row.get("Flow"),
            row.get("Device"),
            row.get("Status"),
        )


def finalize_workbook(excel_path: Path, *, file_lock: threading.Lock | None = None) -> None:
    excel_path = excel_path.resolve()
    if not excel_path.is_file():
        logger.warning("finalize: file missing %s", excel_path)
        return
    lock = file_lock or _lock_for(excel_path)
    with lock:
        wb = load_workbook(excel_path)
        wb.save(excel_path)
    logger.info("Excel finalized: %s", excel_path)
