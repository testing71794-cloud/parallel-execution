"""Excel + summary.json for intelligent layer."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger("intelligent_platform.report")

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")


def generate_report(
    analyzed: list[dict[str, Any]],
    clusters: list[dict[str, Any]],
    output_dir: Path,
    basename: str = "ai_intelligence_report",
) -> tuple[Path, Path]:
    """
    Writes:
    - {basename}.xlsx with Failure Category, Root Cause, Suggestion, Cluster ID, Confidence
    - summary.json aggregate
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / f"{basename}.xlsx"
    json_path = output_dir / "summary.json"

    wb = Workbook()
    ws = wb.active
    ws.title = "Failures"

    headers = [
        "Test Name",
        "Suite",
        "Flow",
        "Device",
        "Failure Category",
        "Root Cause",
        "Suggestion",
        "Cluster ID",
        "Confidence",
        "Test issue?",
        "Log/Artifact source",
        "AI Status",
        "Model Used",
        "Analysis Source",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for ri, row in enumerate(analyzed, start=2):
        ws.cell(row=ri, column=1, value=row.get("test_name", ""))
        ws.cell(row=ri, column=2, value=row.get("suite", ""))
        ws.cell(row=ri, column=3, value=row.get("flow", ""))
        ws.cell(row=ri, column=4, value=row.get("device", ""))
        ws.cell(row=ri, column=5, value=row.get("category", ""))
        ws.cell(row=ri, column=6, value=row.get("root_cause", ""))
        ws.cell(row=ri, column=7, value=row.get("suggestion", ""))
        ws.cell(row=ri, column=8, value=row.get("cluster_id", ""))
        ws.cell(row=ri, column=9, value=row.get("confidence", ""))
        ws.cell(row=ri, column=10, value="yes" if row.get("is_test_issue") else "no")
        ws.cell(row=ri, column=11, value=row.get("source", ""))
        ws.cell(row=ri, column=12, value=row.get("ai_status", ""))
        ws.cell(row=ri, column=13, value=row.get("model_used", ""))
        ws.cell(row=ri, column=14, value=row.get("analysis_source", ""))

    ws2 = wb.create_sheet("Clusters")
    ws2.append(["cluster_id", "root_issue", "count", "affected_tests"])
    for r in ws2[1]:
        r.font = Font(bold=True)
        r.fill = HEADER_FILL
    for c in clusters:
        ws2.append(
            [
                c.get("cluster_id"),
                c.get("root_issue"),
                c.get("count"),
                ", ".join(c.get("affected_tests", [])),
            ]
        )

    wb.save(xlsx_path)
    logger.info("Wrote Excel: %s", xlsx_path)

    summary = {
        "version": 1,
        "failures_analyzed": len(analyzed),
        "cluster_count": len(clusters),
        "clusters": clusters,
        "top_cluster": clusters[0] if clusters else None,
    }
    json_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.info("Wrote summary.json: %s", json_path)
    return xlsx_path, json_path
