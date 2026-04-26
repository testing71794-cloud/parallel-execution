#!/usr/bin/env python3
from __future__ import annotations

import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


def parse_status_file(file_path: Path) -> dict:
    data = {
        "suite": "",
        "flow": "",
        "device": "",
        "status": "",
        "exit_code": "",
        "log": "",
        "file_name": file_path.name,
    }
    for line in file_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        if "=" in line:
            k, v = line.split("=", 1)
            data[k.strip().lower()] = v.strip()
    if not data["flow"]:
        parts = file_path.stem.split("__")
        if len(parts) >= 3:
            data["suite"] = parts[0]
            data["flow"] = parts[1]
            data["device"] = parts[2]
    data["status"] = (data.get("status") or "UNKNOWN").upper()
    return data


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: python scripts/generate_build_summary.py <status_dir> <output_dir>")
        return 1

    status_dir = Path(sys.argv[1]).resolve()
    output_dir = Path(sys.argv[2]).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    results = [parse_status_file(p) for p in sorted(status_dir.glob("*.txt")) if parse_status_file(p).get("status") != "RUNNING"]
    if not results:
        print(f"Warning: no completed status files in {status_dir} — writing empty build summary")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Kodak Smile Execution Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Generated On"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"] = "Total Results"
    ws["B3"] = len(results)
    ws["A4"] = "Passed"
    ws["B4"] = sum(1 for r in results if r.get("status") == "PASS")
    ws["A5"] = "Failed"
    ws["B5"] = sum(1 for r in results if r.get("status") not in ("PASS",))

    start = 8
    headers = ["Suite", "Flow", "Device", "Status", "Exit Code", "Log"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=start, column=idx, value=header).font = Font(bold=True)

    row_idx = start + 1
    for row in results:
        ws.cell(row=row_idx, column=1, value=row.get("suite", ""))
        ws.cell(row=row_idx, column=2, value=row.get("flow", ""))
        ws.cell(row=row_idx, column=3, value=row.get("device", ""))
        ws.cell(row=row_idx, column=4, value=row.get("status", ""))
        ws.cell(row=row_idx, column=5, value=row.get("exit_code", ""))
        ws.cell(row=row_idx, column=6, value=row.get("log", ""))
        row_idx += 1

    suite_counter: defaultdict = defaultdict(Counter)
    for row in results:
        suite_counter[row.get("suite", "")][row.get("status", "UNKNOWN")] += 1

    ws2 = wb.create_sheet("Suite Summary")
    ws2.append(["Suite", "Pass", "Fail", "Total"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for suite in sorted(suite_counter):
        passed = suite_counter[suite]["PASS"]
        total = sum(suite_counter[suite].values())
        failed = total - passed
        ws2.append([suite, passed, failed, total])

    # Do not clobber final_execution_report.xlsx (owned by generate_excel_report.py merge)
    final_xlsx = output_dir / "build_summary_overview.xlsx"
    wb.save(final_xlsx)

    html = [
        "<html><body>",
        "<h2>Kodak Smile Execution Summary</h2>",
        f"<p>Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
        "<table border='1' cellspacing='0' cellpadding='4'>",
        "<tr><th>Suite</th><th>Flow</th><th>Device</th><th>Status</th><th>Exit Code</th></tr>",
    ]
    for row in results:
        html.append(
            f"<tr><td>{row.get('suite','')}</td><td>{row.get('flow','')}</td><td>{row.get('device','')}</td><td>{row.get('status','')}</td><td>{row.get('exit_code','')}</td></tr>"
        )
    html.append("</table></body></html>")
    (output_dir / "summary.html").write_text("\n".join(html), encoding="utf-8")

    print(f"Build summary generated: {final_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
