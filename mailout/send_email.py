"""
Send final_execution_report.xlsx after parallel orchestration completes (HTML + attachments).

HTML body shows FAILED tests only (Suite / Flow / Device / Status / Failure Reason /
AI Analysis / Screenshot / Video), matching the Jenkins Failed Tests report, plus a
Failed Test Artifacts zip link when BUILD_URL is set.

Default attachments:
  1) final_execution_report.xlsx
  2) execution_logs.zip (existing, or auto-built from reports/**/*.log)
  3) failed_tests_artifacts.zip when present (failed logs/screenshots/videos)

Optional AI files (intelligent_platform): set ORCH_EMAIL_ATTACH_AI=1, then
  + ai_intelligence_report.xlsx, intelligence_result.json when present.

Env ORCH_AI_INTELLIGENCE_XLSX / AI_INTELLIGENCE_REPORT_XLSX can override the AI Excel path.
"""
from __future__ import annotations

import html
import logging
import os
import socket
import smtplib
import ssl
import sys
import zipfile
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path

from openpyxl import load_workbook

_REPO = Path(__file__).resolve().parents[1]
if str(_REPO) not in sys.path:
    sys.path.insert(0, str(_REPO))
from utils.device_utils import render_device_display  # noqa: E402
from utils.git_branch import detect_git_branch  # noqa: E402

logger = logging.getLogger("orch.mail")


def resolve_final_excel_path(root: Path) -> Path | None:
    """
    Locate final_execution_report.xlsx for attachment.
    Order: explicit env path → repo root → build-summary → shallowest rglob under root.
    """
    root = root.resolve()
    for env_name in ("FINAL_EXECUTION_REPORT_XLSX", "ORCH_EXCEL_OUT"):
        raw = os.getenv(env_name, "").strip()
        if not raw:
            continue
        p = Path(raw)
        if not p.is_absolute():
            p = (root / p).resolve()
        if p.is_file():
            logger.info("Using Excel from %s=%s", env_name, p)
            return p
        logger.warning("Env %s points to missing file: %s", env_name, p)

    candidates = [
        root / "final_execution_report.xlsx",
        root / "build-summary" / "final_execution_report.xlsx",
    ]
    for c in candidates:
        if c.is_file():
            logger.info("Using Excel: %s", c)
            return c

    matches = [p for p in root.rglob("final_execution_report.xlsx") if p.is_file()]
    if matches:
        best = min(matches, key=lambda p: (len(p.parts), str(p)))
        logger.info("Using Excel (search): %s", best)
        return best

    logger.error(
        "No final_execution_report.xlsx under %s (tried env FINAL_EXECUTION_REPORT_XLSX / ORCH_EXCEL_OUT, root, build-summary, rglob)",
        root,
    )
    return None


def resolve_execution_logs_zip(excel_path: Path, root: Path | None) -> Path | None:
    """Use existing build-summary/execution_logs.zip (or next to the Excel) when present."""
    candidates: list[Path] = [excel_path.parent / "execution_logs.zip"]
    if root is not None:
        r = root.resolve()
        candidates.extend(
            [
                r / "build-summary" / "execution_logs.zip",
                r / "execution_logs.zip",
            ]
        )
    seen: set[Path] = set()
    for p in candidates:
        rp = p.resolve()
        if rp in seen:
            continue
        seen.add(rp)
        if rp.is_file():
            logger.info("Found execution logs zip: %s", rp)
            return rp
    return None


def _collect_log_files_for_zip(root: Path) -> list[Path]:
    """
    All *.log under reports/ and status/ (orchestrator + Maestro logs) for execution_logs.zip.
    """
    r = root.resolve()
    seen: set[Path] = set()
    out: list[Path] = []
    for sub in ("reports", "status", "collected-artifacts"):
        d = r / sub
        if not d.is_dir():
            continue
        for p in d.rglob("*.log"):
            if p.is_file():
                k = p.resolve()
                if k not in seen:
                    seen.add(k)
                    out.append(p)
    return sorted(out, key=lambda p: str(p))


def build_execution_logs_zip(root: Path) -> Path | None:
    """
    Create build-summary/execution_logs.zip from reports/**/*.log when any logs exist.
    """
    r = root.resolve()
    log_files = _collect_log_files_for_zip(r)
    if not log_files:
        return None
    out_dir = r / "build-summary"
    out_dir.mkdir(parents=True, exist_ok=True)
    zpath = out_dir / "execution_logs.zip"
    with zipfile.ZipFile(zpath, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in log_files:
            try:
                arc = p.relative_to(r)
            except ValueError:
                arc = p.name
            zf.write(p, arcname=str(arc).replace("\\", "/"))
    logger.info("Created %s with %d log file(s)", zpath, len(log_files))
    return zpath


def resolve_or_build_execution_logs_zip(excel_path: Path, root: Path | None) -> Path | None:
    """Prefer an existing execution_logs.zip; otherwise zip all reports/**/*.log if any."""
    z = resolve_execution_logs_zip(excel_path, root)
    if z is not None:
        return z
    if root is None:
        return None
    return build_execution_logs_zip(root)


def resolve_ai_intelligence_artifacts(root: Path) -> list[Path]:
    """
    intelligent_platform outputs (when present):
    - ai_intelligence_report.xlsx (AI Analyses workbook)
    - intelligence_result.json (full pipeline result)
    """
    r = root.resolve()
    out: list[Path] = []
    seen: set[Path] = set()

    def _add(p: Path) -> None:
        if not p.is_file():
            return
        key = p.resolve()
        if key in seen:
            return
        seen.add(key)
        out.append(p)

    for env_name in ("ORCH_AI_INTELLIGENCE_XLSX", "AI_INTELLIGENCE_REPORT_XLSX"):
        raw = os.getenv(env_name, "").strip()
        if not raw:
            continue
        p = Path(raw)
        if not p.is_absolute():
            p = (r / p).resolve()
        if p.is_file():
            _add(p)
            break

    if not any(p.suffix.lower() == ".xlsx" for p in out):
        _add(r / "build-summary" / "ai_intelligence_report.xlsx")
    if not any(p.suffix.lower() == ".xlsx" for p in out):
        alts = [p for p in r.rglob("ai_intelligence_report.xlsx") if p.is_file()]
        if alts:
            _add(min(alts, key=lambda p: (len(p.parts), str(p))))

    _add(r / "build-summary" / "intelligence_result.json")
    return out


def _normalize_header(s: str) -> str:
    return str(s or "").strip().lower().replace("  ", " ")


def _row_values_to_len(row, n_cols: int) -> list:
    r = list(row) if row is not None else []
    if len(r) < n_cols:
        r.extend([None] * (n_cols - len(r)))
    return r


def _sheet_headers_from_ws(ws) -> list[str] | None:
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row1:
        return None
    return [str(h or "").strip() for h in row1]


def _sheet_looks_flow_tabular(headers: list[str]) -> bool:
    hlow = { _normalize_header(x) for x in headers if str(x or "").strip() }
    if "status" not in hlow and "test status" not in hlow:
        return False
    if "suite" not in hlow and "test suite" not in hlow and "suite name" not in hlow:
        return False
    return any(
        k in hlow
        for k in (
            "flow",
            "flow name",
            "test name",
        )
    )


def _ordered_tabular_candidate_names(wb) -> list[str]:
    """Sheets that look like execution data, ordered: Flow Report, Raw Results, then others."""
    out: list[str] = []
    for want in ("Flow Report", "Raw Results"):
        if want in wb.sheetnames and want not in out:
            out.append(want)
    for w in wb.sheetnames:
        wstr = str(w or "").strip()
        if wstr.lower() in ("flow report", "raw results"):
            if w not in out:
                out.append(w)
    for w in wb.sheetnames:
        if w in out:
            continue
        h = _sheet_headers_from_ws(wb[w])
        if h and _sheet_looks_flow_tabular(h):
            out.append(w)
    return out


def _col_index(headers: list[str], *candidates: str) -> int | None:
    lower = { _normalize_header(h): i for i, h in enumerate(headers) if str(h or "").strip()}
    for c in candidates:
        key = _normalize_header(c)
        if key in lower:
            return lower[key]
    return None


def _ai_indices_in_order(headers: list[str]) -> list[int]:
    out: list[int] = []
    for cand in (
        "AI Analysis",
        "AI Analyses",
        "AI Failure Summary",
        "AI Status",
        "Root Cause",
        "Suggested Fix",
    ):
        i = _col_index(headers, cand)
        if i is not None and i not in out:
            out.append(i)
    return out


def _ai_cell_to_email(raw: str) -> str:
    t = (raw or "").strip()
    if not t or t in ("N/A", "NOT_CHECKED"):
        return "—"
    return t


def _flow_cell_invalid(flow_raw: str) -> bool:
    t = (flow_raw or "").strip()
    if not t:
        return True
    if t in (
        "—",
        "–",
        "―",
        "-",
        "—",  # unicode
        "n/a",
        "N/A",
    ):
        return True
    return False


def _simplified_resolve_device(disp: str, did: str) -> str:
    """Display-only: map stored serials to friendly names for email tables."""
    return render_device_display(disp, did)


def _apply_display_to_email_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Final render pass: ensure device column shows friendly names, not serials."""
    out: list[dict[str, str]] = []
    for r in rows:
        nr = dict(r)
        nr["device"] = render_device_display(nr.get("device", ""), nr.get("device_id", ""))
        out.append(nr)
    return out


def _git_branch_for_summary(sheet_kv: dict[str, str]) -> str:
    branch = (sheet_kv.get("Git Branch") or "").strip()
    if branch and branch.lower() != "unknown":
        return branch
    return detect_git_branch(_REPO)


def _parse_table_rows_for_sheet(
    headers: list[str], rows_iter, _sheet_title: str = ""
) -> list[dict[str, str]]:
    n_cols = len(headers)

    i_suite = _col_index(
        headers, "Suite", "Test suite", "Suite Name"
    )
    i_flow = _col_index(
        headers,
        "Flow",
        "Flow Name",
        "Test Name",
    )
    i_dname = _col_index(
        headers,
        "Device Name",
        "Device",
    )
    i_did = _col_index(
        headers,
        "Device ID",
        "Device Id",
        "Udid",
        "UDID",
        "Serial",
    )
    i_status = _col_index(headers, "Status", "Test status")
    i_exit = _col_index(headers, "Exit Code", "Exit code", "ExitCode", "exit_code")
    ai_idx_list = _ai_indices_in_order(headers)

    if i_status is None or i_flow is None:
        return []

    out: list[dict[str, str]] = []
    for row in rows_iter:
        if not row and n_cols:
            continue
        cells = _row_values_to_len(row, n_cols)
        if all(v is None or str(v).strip() == "" for v in cells):
            continue

        def _cell(i: int | None) -> str:
            if i is None or i < 0 or i >= len(cells):
                return ""
            v = cells[i]
            return "" if v is None else str(v).strip()

        suite = _cell(i_suite) if i_suite is not None else ""
        flow = _cell(i_flow)
        if _flow_cell_invalid(flow):
            continue

        st = (_cell(i_status) or "").upper() or "UNKNOWN"
        ex = _cell(i_exit)
        if ex == "":
            ex = "0"

        disp = _cell(i_dname) if i_dname is not None else ""
        did = _cell(i_did) if i_did is not None else ""
        device = _simplified_resolve_device(disp, did)

        raw_ai_parts: list[str] = []
        for j in ai_idx_list:
            t = _cell(j)
            if t and t not in raw_ai_parts:
                raw_ai_parts.append(t)
        raw_ai = " | ".join(raw_ai_parts) if raw_ai_parts else ""
        ai = _ai_cell_to_email(raw_ai)

        out.append(
            {
                "suite": suite,
                "flow": flow,
                "device": device,
                "device_id": did,
                "status": st,
                "exit_code": ex,
                "ai_analyses": ai,
            }
        )
    return out


def _drop_unknown_mixed_simpler(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """If any row is PASS/FAIL/FLAKY, drop only rows that are only UNKNOWN and look like spurious rollups."""
    if not rows:
        return rows
    has_pfb = any((r.get("status") or "").upper() in ("PASS", "FAIL", "FLAKY") for r in rows)
    if not has_pfb:
        return rows
    return [r for r in rows if (r.get("status") or "").upper() != "UNKNOWN"]


def read_execution_table_rows(
    excel_path: Path,
) -> tuple[list[dict[str, str]], str | None]:
    """
    Return flow-wise rows: suite, flow, device, status, exit_code, ai_analyses.
    Picks the best data sheet and maps columns robustly. Does not use read_only so
    short rows are padded to the header width (fixes ragged/merged cell reads).
    """
    path = excel_path.resolve()
    _log = logger.info
    _print = print
    _log("[email] final_execution Excel path: %s", path)
    _print(f"[orch.mail] final_execution Excel path: {path}")

    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        names = _ordered_tabular_candidate_names(wb)
        _print(f"[orch.mail] workbook sheets: {list(wb.sheetnames)}")
        _log("Available sheet names: %s", wb.sheetnames)
        if not names:
            return [], "No tabular execution sheet found (need Suite, Flow, Status in header row 1)."

        parsed_triples: list[tuple[str, list[dict[str, str]], list[str]]] = []
        for name in names:
            ws = wb[name]
            h = _sheet_headers_from_ws(ws)
            if not h or not _sheet_looks_flow_tabular(h):
                _log("Skip sheet (header mismatch): %s", name)
                continue
            pr = _parse_table_rows_for_sheet(
                h, ws.iter_rows(min_row=2, values_only=True), name
            )
            n = len(pr)
            _log("Sheet %r: parsed %d row(s) before project-wide filters", name, n)
            _print(
                f"[orch.mail] sheet {name!r}: {n} row(s) parsed; header: {h!r}"
            )
            parsed_triples.append((name, pr, h))

        if not parsed_triples:
            return [], "No execution data rows (after parsing candidate sheets)."

        pref = {"Flow Report": 0, "Raw Results": 1}
        sheet_used, out, headers_used = max(
            parsed_triples,
            key=lambda t: (len(t[1]), -pref.get(t[0], 3)),
        )
        out = _drop_unknown_mixed_simpler(out)
        if not out:
            return (
                [],
                f"All rows on sheet {sheet_used!r} were filtered (empty/invalid flow or UNKNOWN-only).",
            )

        _print(
            f"[orch.mail] selected sheet: {sheet_used!r} | total email rows: {len(out)}"
        )
        _print(f"[orch.mail] detected column headers: {headers_used!r}")
        _log("Selected sheet: %r; detected columns: %s", sheet_used, headers_used)
        _log("Using %d email table row(s)", len(out))
        return out, None
    finally:
        wb.close()


def read_summary_sheet_key_values(excel_path: Path) -> dict[str, str]:
    """Key/value pairs from the 'Summary' sheet (column A = label, B = value), like the Excel preview."""
    out: dict[str, str] = {}
    path = excel_path.resolve()
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Summary" not in wb.sheetnames:
            return out
        ws = wb["Summary"]
        for r in range(1, 20):
            k = ws.cell(r, 1).value
            v = ws.cell(r, 2).value
            if k is None or str(k).strip() == "":
                continue
            key = str(k).strip()
            if not key:
                continue
            val = "" if v is None else str(v).strip()
            if r == 1 and not val and "kodak" in key.lower() and "smile" in key.lower():
                continue
            out[key] = val
    finally:
        wb.close()
    return out


def compute_summary_from_rows(table_rows: list[dict[str, str]]) -> dict[str, str]:
    """When there is no Summary sheet, match merged-report semantics from Raw rows."""
    t = len(table_rows)
    passed = 0
    flaky = 0
    for r in table_rows:
        st = (r.get("status") or "").upper()
        if st == "PASS":
            passed += 1
        elif st == "FLAKY":
            flaky += 1
    non_pass = t - passed
    return {
        "Total rows": str(t),
        "Passed": str(passed),
        "Flaky": str(flaky),
        "Failed (non-PASS)": str(non_pass - flaky),
    }


def build_summary_display_pairs(
    sheet_kv: dict[str, str], table_rows: list[dict[str, str]], generated_on: str
) -> list[tuple[str, str]]:
    """
    Build ordered (label, value) lines for the email, matching the Excel Summary sheet
    (Total / Passed / Failed) like the Gmail xlsx thumbnail.
    """
    if sheet_kv:
        rows_out: list[tuple[str, str]] = []
        if "Total rows" in sheet_kv:
            rows_out.append(("Total tests", sheet_kv["Total rows"]))
        elif "Total" in sheet_kv:
            rows_out.append(("Total tests", sheet_kv["Total"]))

        if "Passed" in sheet_kv:
            rows_out.append(("Passed", sheet_kv["Passed"]))

        fk = "Failed (non-PASS, excl. flaky count below)"
        if fk in sheet_kv:
            rows_out.append(("Failed", sheet_kv[fk]))
        elif "Failed" in sheet_kv and fk not in sheet_kv:
            rows_out.append(("Failed", sheet_kv["Failed"]))

        if "Flaky" in sheet_kv and sheet_kv["Flaky"] not in ("0", ""):
            rows_out.append(("Flaky", sheet_kv["Flaky"]))

        rows_out.append(("Git Branch", _git_branch_for_summary(sheet_kv)))

        if "Generated" in sheet_kv and str(sheet_kv["Generated"]).strip():
            rows_out.append(("Generated on", sheet_kv["Generated"]))
        else:
            rows_out.append(("Generated on", generated_on))
        return rows_out

    comp = compute_summary_from_rows(table_rows)
    rows_out = [
        ("Total tests", comp.get("Total rows", "0")),
        ("Passed", comp.get("Passed", "0")),
        ("Failed", comp.get("Failed (non-PASS)", "0")),
    ]
    if int(comp.get("Flaky", "0") or "0") > 0:
        rows_out.append(("Flaky", comp.get("Flaky", "0")))
    rows_out.append(("Git Branch", _git_branch_for_summary(sheet_kv)))
    rows_out.append(("Generated on", generated_on))
    return rows_out


def _summary_stats_html(pairs: list[tuple[str, str]]) -> str:
    if not pairs:
        return ""
    trs = []
    for label, value in pairs:
        trs.append(
            "<tr>"
            f'<th scope="row" style="text-align:left; padding:6px 10px; border:1px solid #ccc; background:#e8f0f8; font-weight:600; white-space:nowrap;">{html.escape(label)}</th>'
            f'<td style="padding:6px 10px; border:1px solid #ccc;">{html.escape(value)}</td>'
            "</tr>"
        )
    return (
        '<p class="sub" style="margin:12px 0 6px; font-weight:600; color:#1f4e79;">'
        "Run overview</p>"
        f'<table class="sum" role="presentation" style="border-collapse:collapse; max-width:480px; margin-bottom:18px;">{"".join(trs)}</table>'
    )


def _attachments_block_html(attachment_labels: list[str]) -> str:
    if not attachment_labels:
        return ""
    items = "\n    ".join(f"<li>{html.escape(f)}</li>" for f in attachment_labels)
    return f"""<p class="sub"><b>Attachments</b></p>
  <ul style="margin:8px 0 16px; padding-left:20px; color:#333;">
    {items}
  </ul>"""


def _format_ai_for_html(ai_text: str) -> str:
    """Shorter in-cell, full text in title for long AI summaries."""
    t = (ai_text or "").strip() or "—"
    one_line = " ".join(t.split())
    if len(t) > 200:
        short = t[:200] + "…"
        return (
            f'<td class="c-ai" title="{html.escape(one_line, quote=True)}">'
            f"{html.escape(short)}</td>"
        )
    return f'<td class="c-ai">{html.escape(t)}</td>'


def _status_html_class(status: str) -> str:
    u = (status or "").upper()
    if u == "PASS":
        return "st-pass"
    if u in ("FAIL", "PARSE_ERROR", "ERROR"):
        return "st-fail"
    if u == "FLAKY":
        return "st-flaky"
    return "st-other"


def build_email_html(
    rows: list[dict[str, str]],
    generated_on: str,
    error_note: str | None,
    attachment_labels: list[str] | None = None,
    summary_pairs: list[tuple[str, str]] | None = None,
    *,
    artifact_url: str | None = None,
) -> str:
    # Email body shows FAILED tests only (Suite/Flow/Device/Status/Reason/AI/Screenshot/Video).
    failed_rows = _filter_failed_email_rows(rows)
    table_body = _failed_tests_summary_html(failed_rows, artifact_url=artifact_url)
    if error_note:
        table_body = f'<p class="warn">{html.escape(error_note)}</p>{table_body}'

    return f"""\
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {{ font-family: Calibri, "Segoe UI", Arial, sans-serif; font-size: 14px; color: #1a1a1a; }}
  h1 {{ color: #1f4e79; font-size: 20px; margin-bottom: 8px; }}
  .sub {{ color: #666; font-size: 13px; margin-bottom: 16px; }}
  .warn {{ color: #a94442; background: #fbe8e6; padding: 8px; border-radius: 4px; }}
  table.ex {{ border-collapse: collapse; width: 100%; max-width: 1200px; border: 1px solid #000; }}
  table.ex th, table.ex td {{ border: 1px solid #000; padding: 8px 10px; text-align: left; vertical-align: top; word-break: break-word; }}
  .c-ai {{ max-width: 280px; font-size: 13px; line-height: 1.35; color: #222; }}
  table.ex th {{ background: #2e5c8a; color: #fff; font-weight: 600; }}
  .st-pass {{ color: #1b5e20; background: #e8f5e9; font-weight: bold; }}
  .st-fail {{ color: #b71c1c; background: #ffebee; font-weight: bold; }}
  .st-flaky {{ color: #e65100; background: #fff3e0; font-weight: bold; }}
  .st-other {{ color: #333; background: #f5f5f5; font-weight: bold; }}
  a {{ color: #1565c0; }}
</style>
</head>
<body>
  <h1>Kodak Smile Execution Summary</h1>
  {_summary_stats_html(summary_pairs if summary_pairs else [("Generated on", generated_on)])}
  {table_body}
  {_attachments_block_html(attachment_labels or [])}
  <p class="sub" style="margin-top:20px;">This message was sent by Jenkins automation. See the attachment list above.</p>
</body>
</html>"""


def _thead() -> str:
    return (
        "<tr>"
        "<th>Suite</th><th>Flow</th><th>Device</th><th>Status</th>"
        "<th>Failure Reason</th><th>AI Analysis</th><th>Screenshot</th><th>Video</th>"
        "</tr>"
    )


def build_email_plain(
    rows: list[dict[str, str]],
    generated_on: str,
    error_note: str | None,
    attachment_labels: list[str] | None = None,
    summary_pairs: list[tuple[str, str]] | None = None,
    *,
    artifact_url: str | None = None,
) -> str:
    lines = [
        "Kodak Smile Execution Summary",
        "",
    ]
    for label, value in summary_pairs or [("Generated on", generated_on)]:
        lines.append(f"{label}: {value}")
    lines.append("")
    if error_note:
        lines.append(error_note)
        lines.append("")
    failed_rows = _filter_failed_email_rows(rows)
    lines.append(_failed_tests_summary_plain(failed_rows, artifact_url=artifact_url))
    lines.append("")
    lines.append("Attachments:")
    for name in attachment_labels or []:
        lines.append(f"  - {name}")
    lines.append("")
    if _orch_email_attach_ai():
        lines.append(
            "The execution workbook, optional AI analysis files, and the log zip (when present) are attached."
        )
    else:
        lines.append(
            "The execution workbook and execution_logs.zip (when log files are present) are attached."
        )
    return "\n".join(lines)


def _is_failure_status(status: str) -> bool:
    return (status or "").upper() in {"FAIL", "FLAKY", "PARSE_ERROR", "ERROR"}


def _filter_failed_email_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [r for r in rows if _is_failure_status(r.get("status", ""))]


def load_failed_tests_summary(root: Path) -> tuple[list[dict], bool]:
    """Read build-summary/failed_tests_summary.json from collect_failed_artifacts.py."""
    import json

    path = root.resolve() / "build-summary" / "failed_tests_summary.json"
    if not path.is_file():
        return [], False
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        logger.warning("Could not read failed_tests_summary.json: %s", exc)
        return [], True
    rows = data.get("failures") if isinstance(data, dict) else None
    if not isinstance(rows, list):
        return [], True
    return [r for r in rows if isinstance(r, dict)], True


def resolve_failed_tests_artifacts_zip(root: Path) -> Path | None:
    p = root.resolve() / "build-summary" / "failed_tests_artifacts.zip"
    return p if p.is_file() else None


def _jenkins_artifact_url(relative_path: str) -> str | None:
    base = getenv_any("BUILD_URL", "JENKINS_BUILD_URL", default="").rstrip("/")
    if not base:
        return None
    rel = relative_path.lstrip("/").replace("\\", "/")
    return f"{base}/artifact/{rel}"


def _row_device_id(row: dict) -> str:
    did = str(row.get("device_id") or "").strip()
    if did:
        return did
    # Fallback: serial sometimes only lives in the Device column
    return str(row.get("device") or "").strip()


def _enrich_rows_from_failed_summary(
    rows: list[dict[str, str]], failed_summary: list[dict]
) -> list[dict[str, str]]:
    if not failed_summary:
        for row in rows:
            if not row.get("failure_reason"):
                row["failure_reason"] = "MAESTRO_FAILED"
        return rows
    by_key: dict[tuple[str, str, str], dict] = {}
    for item in failed_summary:
        suite = str(item.get("suite") or "").strip().casefold()
        flow = str(item.get("test_name") or item.get("flow") or "").strip()
        dev = str(item.get("device_id") or "").strip()
        by_key[(suite, flow, dev)] = item

    out: list[dict[str, str]] = []
    for row in rows:
        nr = dict(row)
        key = (
            str(nr.get("suite") or "").casefold(),
            str(nr.get("flow") or "").strip(),
            _row_device_id(nr),
        )
        hit = by_key.get(key)
        if not hit:
            # Retry without requiring exact device match when only one summary hit for suite+flow
            soft = [
                v
                for (s, f, _d), v in by_key.items()
                if s == key[0] and f == key[1]
            ]
            if len(soft) == 1:
                hit = soft[0]
        if hit:
            if hit.get("video_artifact"):
                nr["video_artifact"] = str(hit.get("video_artifact") or "")
            if hit.get("screenshot_artifact"):
                nr["screenshot_artifact"] = str(hit.get("screenshot_artifact") or "")
            if hit.get("failure_reason") and not nr.get("failure_reason"):
                nr["failure_reason"] = str(hit.get("failure_reason") or "")
            if hit.get("device_id") and not nr.get("device_id"):
                nr["device_id"] = str(hit.get("device_id") or "")
        if not nr.get("failure_reason"):
            nr["failure_reason"] = "MAESTRO_FAILED"
        out.append(nr)
    return out


def _failed_tests_summary_html(rows: list[dict], *, artifact_url: str | None = None) -> str:
    """Failed-tests-only table matching Step Print / Jenkins email layout."""
    if not rows:
        return (
            '<p class="sub" style="margin:12px 0 16px; font-weight:600; color:#1b5e20;">'
            "No failed tests detected."
            "</p>"
        )
    trs = [
        "<tr>"
        "<th>Suite</th><th>Flow</th><th>Device</th><th>Status</th>"
        "<th>Failure Reason</th><th>AI Analysis</th><th>Screenshot</th><th>Video</th>"
        "</tr>"
    ]
    for row in rows:
        suite = str(row.get("suite") or "—")
        name = str(row.get("test_name") or row.get("flow") or "—")
        device = str(row.get("device") or row.get("device_id") or "—")
        status = str(row.get("status") or "FAIL")
        reason = str(
            row.get("failure_reason") or row.get("reason") or row.get("error_message") or "MAESTRO_FAILED"
        )
        ai = str(row.get("ai_analyses") or row.get("ai_analysis") or "—").strip() or "—"
        shot = str(row.get("screenshot_artifact") or "").strip()
        video = str(row.get("video_artifact") or "").strip()
        shot_cell = html.escape(shot) if shot else "—"
        video_cell = html.escape(video) if video else "—"
        if shot:
            shot_url = _jenkins_artifact_url(f"build-summary/failed-artifacts/{shot}")
            if shot_url:
                shot_cell = f'<a href="{html.escape(shot_url)}">{html.escape(shot)}</a>'
        if video:
            video_url = _jenkins_artifact_url(f"build-summary/failed-artifacts/{video}")
            if video_url:
                video_cell = f'<a href="{html.escape(video_url)}">{html.escape(video)}</a>'
        cls = _status_html_class(status)
        ai_short = ai if len(ai) <= 200 else ai[:200] + "…"
        trs.append(
            "<tr>"
            f"<td>{html.escape(suite)}</td>"
            f"<td>{html.escape(name)}</td>"
            f"<td>{html.escape(device)}</td>"
            f'<td class="{cls}"><strong>{html.escape(status)}</strong></td>'
            f"<td>{html.escape(reason)}</td>"
            f'<td class="c-ai" title="{html.escape(ai, quote=True)}">{html.escape(ai_short)}</td>'
            f"<td>{shot_cell}</td>"
            f"<td>{video_cell}</td>"
            "</tr>"
        )
    block = (
        '<p class="sub" style="margin:12px 0 6px; font-weight:600; color:#1f4e79;">'
        "Failed Tests</p>"
        '<table class="ex" role="presentation" style="margin-bottom:18px;">'
        f'{"".join(trs)}</table>'
    )
    if artifact_url:
        block += (
            '<p class="sub"><b>Failed Test Artifacts:</b> '
            f'<a href="{html.escape(artifact_url)}">{html.escape(artifact_url)}</a></p>'
        )
    else:
        block += (
            '<p class="sub"><b>Failed Test Artifacts:</b> '
            "<code>build-summary/failed_tests_artifacts.zip</code> "
            "(attached when available; set BUILD_URL on Jenkins for a clickable link)</p>"
        )
    return block


def _failed_tests_summary_plain(rows: list[dict], *, artifact_url: str | None = None) -> str:
    lines = [
        "Failed Tests",
        "Suite | Flow | Device | Status | Failure Reason | AI Analysis | Screenshot | Video",
        "-" * 100,
    ]
    if not rows:
        lines.append("No failed tests detected.")
    else:
        for row in rows:
            suite = str(row.get("suite") or "—")
            name = str(row.get("test_name") or row.get("flow") or "—")
            device = str(row.get("device") or row.get("device_id") or "—")
            status = str(row.get("status") or "FAIL")
            reason = str(row.get("failure_reason") or row.get("reason") or "MAESTRO_FAILED")
            ai = str(row.get("ai_analyses") or row.get("ai_analysis") or "—")[:80]
            shot = str(row.get("screenshot_artifact") or "").strip() or "—"
            video = str(row.get("video_artifact") or "").strip() or "—"
            lines.append(
                f"{suite} | {name} | {device} | {status} | {reason} | {ai} | {shot} | {video}"
            )
    if artifact_url:
        lines.append(f"Failed Test Artifacts: {artifact_url}")
    return "\n".join(lines)


def getenv_any(*names: str, default: str = "") -> str:
    for name in names:
        v = os.getenv(name, "").strip()
        if v:
            return v
    return default


def _add_file_attachment(msg: EmailMessage, path: Path) -> None:
    data = path.read_bytes()
    name = path.name
    ext = path.suffix.lower()
    if ext == ".json":
        msg.add_attachment(data, maintype="application", subtype="json", filename=name)
    elif ext == ".xlsx":
        msg.add_attachment(
            data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=name,
        )
    elif ext == ".zip":
        msg.add_attachment(data, maintype="application", subtype="zip", filename=name)
    else:
        msg.add_attachment(data, maintype="application", subtype="octet-stream", filename=name)
    logger.info("Attached: %s", name)


def _orch_email_attach_ai() -> bool:
    """Set ORCH_EMAIL_ATTACH_AI=1 to add ai_intelligence_report.xlsx + intelligence_result.json."""
    return getenv_any("ORCH_EMAIL_ATTACH_AI", default="").lower() in ("1", "true", "yes", "on")


def _orch_email_strict() -> bool:
    return (os.environ.get("ORCH_EMAIL_STRICT", "").strip().lower() in ("1", "true", "yes", "on"))


def _smtp_config_ready() -> bool:
    """
    True when user/pass/receiver and server (or implied Gmail) are all present to attempt SMTP.
    """
    smtp_user = getenv_any("SMTP_USER", "SENDER_EMAIL", "GMAIL_USER")
    smtp_server = getenv_any("SMTP_SERVER", "SMTP_HOST")
    if not smtp_server and smtp_user and "@" in smtp_user and smtp_user.lower().rstrip().endswith(
        ("@gmail.com", "@googlemail.com")
    ):
        smtp_server = "smtp.gmail.com"
    smtp_pass = getenv_any("SMTP_PASS", "SMTP_PASSWORD", "SENDER_PASSWORD", "GMAIL_APP_PASSWORD")
    receiver = getenv_any("RECEIVER_EMAIL", "MAIL_TO", "EMAIL_RECIPIENTS", "RECIPIENT", "ORCH_MAIL_TO")
    return bool(smtp_server and smtp_user and smtp_pass and receiver)


def send_execution_report_email(
    excel_path: Path,
    *,
    root: Path | None = None,
    subject: str | None = None,
    body: str | None = None,
) -> bool:
    """
    Returns True if mail was sent, False if skipped or failed (logged).
    HTML body is built from final_execution_report.xlsx unless ``body`` is set (overrides).
    """
    excel_path = excel_path.resolve()
    if not excel_path.is_file():
        logger.error("Excel attachment missing: %s", excel_path)
        return False

    smtp_user = getenv_any("SMTP_USER", "SENDER_EMAIL", "GMAIL_USER")
    smtp_server = getenv_any("SMTP_SERVER", "SMTP_HOST")
    if not smtp_server and smtp_user and "@" in smtp_user and smtp_user.lower().rstrip().endswith(
        ("@gmail.com", "@googlemail.com")
    ):
        smtp_server = "smtp.gmail.com"
        logger.info("Defaulting SMTP server to smtp.gmail.com (Gmail sender)")

    smtp_port_raw = getenv_any("SMTP_PORT", default="587")
    smtp_pass = getenv_any(
        "SMTP_PASS",
        "SMTP_PASSWORD",
        "SENDER_PASSWORD",
        "GMAIL_APP_PASSWORD",
    )
    sender = getenv_any("SENDER_EMAIL", "SMTP_USER", "GMAIL_USER")
    receiver = getenv_any(
        "RECEIVER_EMAIL",
        "MAIL_TO",
        "EMAIL_RECIPIENTS",
        "RECIPIENT",
        "ORCH_MAIL_TO",
    )

    subj = subject or getenv_any(
        "ORCH_MAIL_SUBJECT",
        default="Jenkins Execution Report",
    )

    gen_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rroot = root.resolve() if root is not None else None
    exc_resolved = excel_path.resolve()

    failed_summary_rows: list[dict] = []
    failed_zip: Path | None = None
    artifact_url: str | None = None
    if rroot is not None:
        failed_summary_rows, _ = load_failed_tests_summary(rroot)
        failed_zip = resolve_failed_tests_artifacts_zip(rroot)
        if failed_zip is not None:
            artifact_url = _jenkins_artifact_url("build-summary/failed_tests_artifacts.zip")

    ai_to_attach: list[Path] = (
        [p for p in resolve_ai_intelligence_artifacts(rroot) if p.resolve() != exc_resolved]
        if rroot is not None and _orch_email_attach_ai()
        else []
    )
    attachment_labels: list[str] = [excel_path.name]
    for ap in ai_to_attach:
        if ap.suffix.lower() == ".xlsx" and "ai_intelligence" in ap.name.lower():
            attachment_labels.append(f"{ap.name} (AI Analyses)")
        elif ap.suffix.lower() == ".json":
            attachment_labels.append(f"{ap.name} (AI analyses - full result)")
        else:
            attachment_labels.append(ap.name)
    logs_zip: Path | None = resolve_or_build_execution_logs_zip(
        excel_path, rroot if rroot is not None else None
    )
    if logs_zip is not None:
        attachment_labels.append(f"{logs_zip.name} (execution logs)")
    if failed_zip is not None:
        attachment_labels.append(
            f"{failed_zip.name} (failed test logs, screenshots, videos)"
        )

    if body is not None:
        text_body = body
        html_body = f"<html><body><pre>{html.escape(body)}</pre></body></html>"
    else:
        table_rows, table_err = read_execution_table_rows(excel_path)
        table_rows = _apply_display_to_email_rows(table_rows)
        # Email body: FAILED tests only (Suite/Flow/Device/Status/Reason/AI/Screenshot/Video)
        failed_email_rows = _filter_failed_email_rows(table_rows)
        failed_email_rows = _enrich_rows_from_failed_summary(
            failed_email_rows, failed_summary_rows
        )
        # Include failures present in collect_failed_artifacts summary but missing from Excel
        seen = {
            (
                str(r.get("suite") or "").casefold(),
                str(r.get("flow") or "").strip(),
                _row_device_id(r),
            )
            for r in failed_email_rows
        }
        for item in failed_summary_rows:
            key = (
                str(item.get("suite") or "").casefold(),
                str(item.get("test_name") or item.get("flow") or "").strip(),
                str(item.get("device_id") or "").strip(),
            )
            if key in seen:
                continue
            seen.add(key)
            failed_email_rows.append(
                {
                    "suite": str(item.get("suite") or ""),
                    "flow": str(item.get("test_name") or item.get("flow") or ""),
                    "device": render_device_display("", str(item.get("device_id") or "")),
                    "device_id": str(item.get("device_id") or ""),
                    "status": str(item.get("status") or "FAIL"),
                    "failure_reason": str(item.get("failure_reason") or "MAESTRO_FAILED"),
                    "ai_analyses": "—",
                    "video_artifact": str(item.get("video_artifact") or ""),
                    "screenshot_artifact": str(item.get("screenshot_artifact") or ""),
                }
            )
        sheet_kv = read_summary_sheet_key_values(excel_path)
        summary_pairs = build_summary_display_pairs(sheet_kv, table_rows, gen_ts)
        text_body = build_email_plain(
            failed_email_rows,
            gen_ts,
            table_err,
            attachment_labels,
            summary_pairs,
            artifact_url=artifact_url,
        )
        html_body = build_email_html(
            failed_email_rows,
            gen_ts,
            table_err,
            attachment_labels,
            summary_pairs,
            artifact_url=artifact_url,
        )

    if not smtp_server or not smtp_user or not smtp_pass or not receiver:
        logger.warning(
            "Email skipped: set at minimum SMTP_USER (or SENDER_EMAIL), "
            "SMTP_PASS (or SMTP_PASSWORD), RECEIVER_EMAIL (or MAIL_TO), and "
            "SMTP_SERVER (or omit for @gmail.com to default to smtp.gmail.com). "
            "Gmail: use an App Password, not the normal account password."
        )
        return False

    try:
        port = int(smtp_port_raw)
    except ValueError:
        port = 587

    msg = EmailMessage()
    msg["Subject"] = subj
    msg["From"] = sender or smtp_user
    msg["To"] = receiver
    msg.set_content(text_body)
    msg.add_alternative(html_body, subtype="html")

    _add_file_attachment(msg, excel_path)
    for ap in ai_to_attach:
        _add_file_attachment(msg, ap)
    if logs_zip is not None:
        _add_file_attachment(msg, logs_zip)
    if failed_zip is not None:
        _add_file_attachment(msg, failed_zip)

    context = ssl.create_default_context()
    try:
        with smtplib.SMTP(smtp_server, port, timeout=60) as server:
            server.starttls(context=context)
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        logger.info("Email sent to %s (HTML + attachments)", receiver)
        return True
    except Exception as e:
        logger.error("Email failed: %s", e)
        # Windows: [Errno 11001] getaddrinfo failed — SMTP host DNS resolution failed on the agent.
        if isinstance(e, socket.gaierror) or (
            isinstance(e, OSError)
            and (getattr(e, "errno", None) == 11001 or getattr(e, "winerror", None) == 11001)
        ):
            logger.error(
                "SMTP DNS/network: host %r port %s could not be resolved or reached from this machine. "
                "On the Jenkins agent: verify DNS (e.g. nslookup %s), outbound firewall for TCP %s, "
                "and proxy/VPN rules. Excel and attachments were prepared; only SMTP send failed.",
                smtp_server,
                port,
                smtp_server,
                port,
            )
        return False


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    root = Path(os.environ.get("WORKSPACE", os.getcwd())).resolve()
    excel = resolve_final_excel_path(root)
    if excel is None:
        return 1
    if not _smtp_config_ready():
        if _orch_email_strict():
            logger.error(
                "ORCH_EMAIL_STRICT=1: SMTP is not fully configured. "
                "Jenkins: in the Send Final Email stage set SMTP_USER, SMTP_PASS, RECEIVER_EMAIL (e.g. in the batch step), "
                "or export them on the agent before: python mailout/send_email.py"
            )
            return 1
        logger.warning(
            "SMTP not configured (set SMTP_USER + SMTP_PASS + RECEIVER_EMAIL, and SMTP_SERVER or @gmail.com); "
            "exiting 0. Set ORCH_EMAIL_STRICT=1 to fail this step if mail is required."
        )
        return 0
    return 0 if send_execution_report_email(excel, root=root) else 1


if __name__ == "__main__":
    raise SystemExit(main())
